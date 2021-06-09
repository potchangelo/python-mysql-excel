# Import ข้อมูลสินค้าและประเภทสินค้าจากไฟล์ Excel (.xlsx) เข้าสู่ Database MySQL

import mysql.connector
from openpyxl import load_workbook

# Excel
# - โหลดไฟล์ และโหลดชีทที่เปิดอยู่
workbook = load_workbook('imported_02.xlsx')
sheet = workbook.active

# Database
# - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
db = mysql.connector.connect(
    host="localhost",
    port=3306,
    user="root",
    password="password1234",
    database='golf_want_to_buy'
)
cursor = db.cursor()

# Let's go
# - โหลดข้อมูลประเภทสินค้าทั้งหมด
sql_select_categories = '''
    SELECT *
    FROM categories
'''
cursor.execute(sql_select_categories)
categories = cursor.fetchall()

# - เปรียบเทียบข้อมูลประเภทสินค้า
# - อันไหนยังไม่มีใน Database ให้เพิ่มลงไปใน List
categories_values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    is_new = True
    category = row[3]

    for c in categories:
        if category == c[1]:
            is_new = False
            break

    if is_new:
        print((category, ))
        categories_values.append((category, ))

# - เพิ่มประเภทสินค้าใหม่ลง Database (ถ้ามี)
if len(categories_values) > 0:
    sql_insert_categories = '''
        INSERT INTO categories (title)
        VALUES (%s)
    '''
    cursor.executemany(sql_insert_categories, categories_values)
    db.commit()
    print('เพิ่มประเภทสินค้าจำนวน ' + str(cursor.rowcount) + ' แถว')
else:
    print('ไม่มีประเภทสินค้าใหม่มาเพิ่ม')

# - โหลดข้อมูลประเภทสินค้าทั้งหมด อีกครั้ง
cursor.execute(sql_select_categories)
categories = cursor.fetchall()

# - เชื่อมต่อ category_id กับสินค้าใหม่ของเรา แล้วเพิ่มลงไปใน List
products_values = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    category_title = row[3]
    category_id = None

    for c in categories:
        if category_title == c[1]:
            category_id = c[0]
            break

    product = (row[0], row[1], row[2], category_id)
    print(product)
    products_values.append(product)

# - เพิ่มสินค้าลง Database
sql_insert_products = '''
    INSERT INTO products (title, price, is_necessary, category_id)
    VALUES (%s, %s, %s, %s);
'''
cursor.executemany(sql_insert_products, products_values)
db.commit()
print('เพิ่มสินค้าจำนวน ' + str(cursor.rowcount) + ' แถว')

# ปิดการเชื่อมต่อ Database
cursor.close()
db.close()