# Import ข้อมูลจากไฟล์ Excel (.xlsx) เข้าสู่ Database MySQL
# ทำเหมือน "import_data_02" แต่คราวนี้ในไฟล์จะไม่ใส่เลข 0 สำหรับสินค้าที่ไม่จำเป็น

import mysql.connector
from openpyxl import load_workbook

def run():
    # Excel
    # - โหลดไฟล์ และโหลดชีทที่เปิดอยู่
    workbook = load_workbook('./files/imported_03.xlsx')
    sheet = workbook.active

    # - ต้องทำการตรวจสอบและแปลงข้อมูลในแต่ละแถวก่อน
    # - จากนั้นค่อยเก็บข้อมูลจากแถวที่ข้อมูลครบถ้วนไว้ใน List
    # - เริ่มต้นจากแถวที่ 2 ไปจนถึงแถวสุดท้าย
    values = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        p = list(row[:3])
        if p[0] is None or p[1] is None:
            continue

        if p[2] is None:
            p[2] = 0

        print(p)
        values.append(p)

    # Database
    # - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
    db = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="password1234",
        database='golf_want_to_buy'
    )

    # - ส่งคำสั่ง SQL ไปให้ MySQL ทำการเพิ่มข้อมูล
    # - ใช้ executemany() เพื่อเพิ่มข้อมูลหลายอัน
    cursor = db.cursor()
    sql = '''
        INSERT INTO products (title, price, is_necessary)
        VALUES (%s, %s, %s);
    '''
    cursor.executemany(sql, values)
    db.commit()

    # - สรุปจำนวนข้อมูลที่เพิ่มไป
    print('เพิ่มข้อมูลจำนวน ' + str(cursor.rowcount) + ' แถว')
