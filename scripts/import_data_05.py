# Import ข้อมูลจากไฟล์ Excel (.xlsx) เข้าสู่ Database MySQL
# เป็นการ Import ข้อมูลทุกแถวมาใส่ใน products
# แต่จะเอาข้อมูลโน้ตสินค้า แยกไปใส่ใน product_notes
# และที่สำคัญ ต้องเชื่อม products และ product_notes เข้ากันให้เรียบร้อย

import mysql.connector
from openpyxl import load_workbook

def run():
    # Excel
    # - โหลดไฟล์ และโหลดชีทที่เปิดอยู่
    workbook = load_workbook('./files/imported_05.xlsx')
    sheet = workbook.active

    # Database
    # - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
    db = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="password1234",
        database='golf_want_to_buy'
    )
    cursor = db.cursor()

    # - เอาเฉพาะข้อมูลสินค้ามาใส่ใน List
    new_products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product = (row[0], row[1], row[2])
        print(product)
        new_products.append(product)

    # - เพิ่มข้อมูลสินค้า
    sql_insert_products = '''
        INSERT INTO products (title, price, is_necessary)
        VALUES (%s, %s, %s);
    '''
    cursor.executemany(sql_insert_products, new_products)
    db.commit()
    print('เพิ่มสินค้าจำนวน ' + str(cursor.rowcount) + ' แถว')

    # - ดึง ID ของสินค้าแรกที่เพิ่มไปเมื่อกี๊ มาเตรียมตัวใช้
    first_product_id = cursor.lastrowid

    # - เอาเฉพาะข้อมูลโน้ตสินค้ามาใส่ใน List และผูก id สินค้าด้วย
    new_product_notes = []
    for (product_id, row) in enumerate(sheet.iter_rows(min_row=2, values_only=True), first_product_id):
        note = row[3]
        if note is not None:
            product_note = (note, product_id)
            print(product_note)
            new_product_notes.append(product_note)

    # - เพิ่มข้อมูลโน้ตสินค้า
    sql_insert_product_notes = '''
        INSERT INTO product_notes (note, product_id)
        VALUES (%s, %s);
    '''
    cursor.executemany(sql_insert_product_notes, new_product_notes)
    db.commit()
    print('เพิ่มโน้ตสินค้าจำนวน ' + str(cursor.rowcount) + ' แถว')

    # ปิดการเชื่อมต่อ Database
    cursor.close()
    db.close()
