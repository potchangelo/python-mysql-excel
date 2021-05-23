# Export ข้อมูลจาก Database MySQL ออกมาเป็นไฟล์ Excel (.xlsx)
# เป็นการ Export ข้อมูลทุกแถว
# เอาข้อมูลชื่อสินค้า, ราคาสินค้า, วันที่บันทึก มาแสดง
# แสดงวันที่บันทึก ในรูปแบบ "25 November 2021"
# จัดเรียงข้อมูลตามราคา จากมากไปน้อย

import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from openpyxl.styles.fills import PatternFill, FILL_SOLID
from openpyxl.styles.borders import Border, Side, BORDER_THIN

def run():
    # Database
    # - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
    db = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="password1234",
        database='golf_want_to_buy'
    )

    # - ส่งคำสั่ง SQL ไปให้ MySQL ทำการโหลดข้อมูล
    # - Python จะรับข้อมูลทั้งหมดมาเป็น List ผ่านคำสั่ง fetchall()
    cursor = db.cursor()
    sql = '''
        SELECT title, price, created_at 
        FROM products 
        ORDER BY price DESC;
    '''
    cursor.execute(sql)
    products = cursor.fetchall()

    # Excel
    # - การตกแต่งทั้งหมด
    border = Border(
        top=Side(border_style=BORDER_THIN, color='333333'),
        right=Side(border_style=BORDER_THIN, color='333333'),
        bottom=Side(border_style=BORDER_THIN, color='333333'),
        left=Side(border_style=BORDER_THIN, color='333333')
    )

    header_style = NamedStyle(name='header')
    header_style.fill = PatternFill(fill_type=FILL_SOLID, fgColor='EEEEEE')
    header_style.border = border
    header_style.font = Font(name='Helvetica Neue', bold=True, size=16)
    header_style.alignment = Alignment(vertical='center')

    row_style = NamedStyle(name='row')
    row_style.border = border
    row_style.font = Font(name='Helvetica Neue', size=16)
    row_style.alignment = Alignment(vertical='center')

    # - สร้างไฟล์ใหม่ สร้างชีท และใส่แถวสำหรับเป็นหัวข้อตาราง
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['ชื่อสินค้า', 'ราคา', 'วันที่บันทึก'])

    # - ตกแต่งโดยรวม
    sheet.row_dimensions[1].height = 32
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 20

    # - ตกแต่งแถวแรก (Header)
    header_row = sheet[1]
    for cell in header_row:
        cell.style = header_style

    # - ใส่ข้อมูลทีละอัน เพิ่มลงไปทีละแถว
    for p in products:
        p_real = list(p)
        p_real[2] = p_real[2].strftime('%d %B %Y')
        print(p_real)
        sheet.append(p_real)

    # - ตกแต่งแถวข้อมูล (Row)
    for (number, row) in enumerate(sheet.iter_rows(min_row=2), 2):
        sheet.row_dimensions[number].height = 32
        for cell in row:
            cell.style = row_style

    # - Export ไฟล์ Excel
    workbook.save(filename="./files/exported_05.xlsx")