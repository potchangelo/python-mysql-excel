# Import ข้อมูลจากไฟล์ Excel (.xlsx) เข้าสู่ Database MySQL
# เป็นการ Import ข้อมูลทุกแถว

import mysql.connector
from openpyxl import load_workbook

def run():
    # Excel
    # - โหลดไฟล์ และโหลดชีทที่เปิดอยู่
    workbook = load_workbook('./files/imported_01.xlsx')
    sheet = workbook.active

    # - เก็บข้อมูล (values_only คือ แบบดิบๆ) ทีละแถวไว้ใน List
    # - เริ่มต้นจากแถวที่ 2 ไปจนถึงแถวสุดท้าย
    values = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)
        values.append(row)

    # Database
    # - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
    db = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="password1234",
        database='golf_want_to_buy'
    )

    # - ส่งคำสั่ง SQL ไปให้ MySQL ทำการเพิ่มข้อมูล
    # - ใช้ executemany() เพื่อเพิ่มข้อมูลหลายอัน
    cursor = db.cursor()
    sql = '''
        INSERT INTO products (title, price, is_necessary)
        VALUES (%s, %s, %s);
    '''
    cursor.executemany(sql, values)
    db.commit()

    # - สรุปจำนวนข้อมูลที่เพิ่มไป
    print('เพิ่มข้อมูลจำนวน ' + str(cursor.rowcount) + ' แถว')

    # ปิดการเชื่อมต่อ Database
    cursor.close()
    db.close()
