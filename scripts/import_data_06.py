# Import ข้อมูลจากไฟล์ Excel (.xlsx) เข้าสู่ Database MySQL
# เป็นการ Import ข้อมูลทุกแถวมาใส่ใน products
# แต่จะเอาข้อมูลแฮชแท็ก แยกไปใส่ใน hashtags (ถ้ายังไม่มี)
# และที่สำคัญ ต้องเชื่อม products และ hashtags เข้ากันให้เรียบร้อย

import mysql.connector
from openpyxl import load_workbook

def run():
    # Excel
    # - โหลดไฟล์ และโหลดชีทที่เปิดอยู่
    workbook = load_workbook('./files/imported_06.xlsx')
    sheet = workbook.active

    # Database
    # - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
    db = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="password1234",
        database='golf_want_to_buy'
    )
    cursor = db.cursor()

    # - โหลดข้อมูลแฮชแท็กทั้งหมด
    sql_select_hashtags = '''
        SELECT * 
        FROM hashtags;
    '''
    cursor.execute(sql_select_hashtags)
    db_hashtags = cursor.fetchall()

    # - เอาแฮชแท็กใหม่ มาใส่ใน List
    new_hashtags = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        hashtags_text = row[3]
        hashtags = hashtags_text.split(' ')

        for hashtag in hashtags:
            is_new = True
            for db_hashtag in db_hashtags:
                if hashtag == db_hashtag[1]:
                    is_new = False
                    break
            for new_hashtag in new_hashtags:
                if hashtag == new_hashtag[0]:
                    is_new = False
                    break

            if is_new:
                print((hashtag,))
                new_hashtags.append((hashtag,))

    # - เพิ่มข้อมูลแฮชแท็ก (ถ้ามีอันใหม่)
    if len(new_hashtags) > 0:
        sql_insert_hashtags = '''
            INSERT INTO hashtags (title)
            VALUES (%s);
        '''
        cursor.executemany(sql_insert_hashtags, new_hashtags)
        db.commit()
        print('เพิ่มแฮชแท็กจำนวน ' + str(cursor.rowcount) + ' แถว')
    else:
        print('ไม่มีแฮชแท็กใหม่มาเพิ่ม')

    # - ใส่สินค้าทั้งหมดลง List
    new_products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product = (row[0], row[1], row[2])
        print(product)
        new_products.append(product)

    # - เพิ่มข้อมูลสินค้า
    sql_insert_products = '''
        INSERT INTO products (title, price, is_necessary)
        VALUES (%s, %s, %s);
    '''
    cursor.executemany(sql_insert_products, new_products)
    db.commit()
    print('เพิ่มสินค้าจำนวน ' + str(cursor.rowcount) + ' แถว')

    # - ดึง ID ของสินค้าแรกที่เพิ่มไปเมื่อกี๊ มาเตรียมตัวใช้
    first_product_id = cursor.lastrowid

    # - โหลดข้อมูลแฮชแท็กทั้งหมด (หลังบันทึกล่าสุดไปแล้ว)
    cursor.execute(sql_select_hashtags)
    db_hashtags = cursor.fetchall()

    # - เชื่อมสินค้ากับแฮชแท็กเข้าด้วยกัน
    new_products_hashtags = []
    for (product_id, row) in enumerate(sheet.iter_rows(min_row=2, values_only=True), first_product_id):
        hashtags_text = row[3]
        hashtags = hashtags_text.split(' ')

        for hashtag in hashtags:
            hashtag_id = None
            for db_hashtag in db_hashtags:
                if hashtag == db_hashtag[1]:
                    hashtag_id = db_hashtag[0]
                    break
            print((product_id, hashtag_id))
            new_products_hashtags.append((product_id, hashtag_id))

    # - ใส่ข้อมูลลงในตารางเชื่อมโยง products_hashtags
    sql = '''
        INSERT INTO products_hashtags (product_id, hashtag_id)
        VALUES (%s, %s);
    '''
    cursor.executemany(sql, new_products_hashtags)
    db.commit()
    print('เพิ่มการเชื่อมโยงจำนวน ' + str(cursor.rowcount) + ' แถว')

    # ปิดการเชื่อมต่อ Database
    cursor.close()
    db.close()
