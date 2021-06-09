# Import ข้อมูลจากไฟล์ Excel (.xlsx) เข้าสู่ Database MySQL
# เป็นการ Import ข้อมูลทุกแถวมาใส่ใน products
# แต่จะเอาข้อมูลประเภทสินค้า แยกไปใส่ใน categories (ถ้ายังไม่มี)
# และที่สำคัญ ต้องเชื่อม products และ categories เข้ากันให้เรียบร้อย

import mysql.connector
from openpyxl import load_workbook

def run():
    # Excel
    # - โหลดไฟล์ และโหลดชีทที่เปิดอยู่
    workbook = load_workbook('./files/imported_04.xlsx')
    sheet = workbook.active

    # Database
    # - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
    db = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="password1234",
        database='golf_want_to_buy'
    )
    cursor = db.cursor()

    # - โหลดข้อมูลประเภทสินค้าทั้งหมด
    sql_select_categories = '''
        SELECT * 
        FROM categories;
    '''
    cursor.execute(sql_select_categories)
    db_categories = cursor.fetchall()

    # - เอาประเภทสินค้าใหม่ มาใส่ใน List
    new_categories = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        is_new = True
        category = row[3]

        for db_category in db_categories:
            if category == db_category[1]:
                is_new = False
                break

        if is_new:
            print((category,))
            new_categories.append((category,))

    # - เพิ่มข้อมูลประเภทสินค้า (ถ้ามีอันใหม่)
    if len(new_categories) > 0:
        sql_insert_categories = '''
            INSERT INTO categories (title)
            VALUES (%s);
        '''
        cursor.executemany(sql_insert_categories, new_categories)
        db.commit()
        print('เพิ่มประเภทสินค้าจำนวน ' + str(cursor.rowcount) + ' แถว')
    else:
        print('ไม่มีประเภทสินค้าใหม่มาเพิ่ม')

    # - โหลดข้อมูลประเภทสินค้าทั้งหมด (หลังบันทึกล่าสุดไปแล้ว)
    cursor.execute(sql_select_categories)
    db_categories = cursor.fetchall()

    # - ใส่ category_id ให้สินค้าแต่ละอย่าง
    new_products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        category_title = row[3]
        category_id = None

        for db_category in db_categories:
            if category_title == db_category[1]:
                category_id = db_category[0]
                break

        product = (row[0], row[1], row[2], category_id)
        print(product)
        new_products.append(product)

    # - เพิ่มข้อมูลสินค้า
    sql_insert_products = '''
        INSERT INTO products (title, price, is_necessary, category_id)
        VALUES (%s, %s, %s, %s);
    '''
    cursor.executemany(sql_insert_products, new_products)
    db.commit()
    print('เพิ่มสินค้าจำนวน ' + str(cursor.rowcount) + ' แถว')

    # ปิดการเชื่อมต่อ Database
    cursor.close()
    db.close()
