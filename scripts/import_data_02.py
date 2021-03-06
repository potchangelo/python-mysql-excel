# Import ข้อมูลจากไฟล์ Excel (.xlsx) เข้าสู่ Database MySQL
# เป็นการ Import เฉพาะแถวที่มีข้อมูลครบถ้วนเท่านั้น

import mysql.connector
from openpyxl import load_workbook

def run():
    # Excel
    # - โหลดไฟล์ และโหลดชีทที่เปิดอยู่
    workbook = load_workbook('./files/imported_02.xlsx')
    sheet = workbook.active

    # - ต้องทำการตรวจสอบข้อมูลในแต่ละแถวก่อน
    # - จากนั้นค่อยเก็บข้อมูลจากแถวที่ข้อมูลครบถ้วนไว้ใน List
    # - เริ่มต้นจากแถวที่ 2 ไปจนถึงแถวสุดท้าย
    values = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        p = row[:3]
        if p[0] is None or p[1] is None or p[2] is None:
            continue
        print(p)
        values.append(p)

    # Database
    # - เชื่อมต่อ Database (เปลี่ยนค่า Connection เป็นของเครื่องตัวเองเน่อ)
    db = mysql.connector.connect(
        host="localhost",
        port=3306,
        user="root",
        password="password1234",
        database='golf_want_to_buy'
    )

    # - ส่งคำสั่ง SQL ไปให้ MySQL ทำการเพิ่มข้อมูล
    # - ใช้ executemany() เพื่อเพิ่มข้อมูลหลายอัน
    cursor = db.cursor()
    sql = '''
        INSERT INTO products (title, price, is_necessary)
        VALUES (%s, %s, %s);
    '''
    cursor.executemany(sql, values)
    db.commit()

    # - สรุปจำนวนข้อมูลที่เพิ่มไป
    print('เพิ่มข้อมูลจำนวน ' + str(cursor.rowcount) + ' แถว')

    # ปิดการเชื่อมต่อ Database
    cursor.close()
    db.close()
